# main_window.py - –≥–ª–∞–≤–Ω–æ–µ –æ–∫–Ω–æ –ø—Ä–∏–ª–æ–∂–µ–Ω–∏—è (–∏—Å–ø—Ä–∞–≤–ª–µ–Ω–Ω–∞—è –≤–µ—Ä—Å–∏—è —Å –ø—Ä–∞–≤–∏–ª—å–Ω—ã–º–∏ –æ—Ç—Å—Ç—É–ø–∞–º–∏)
import os
import sys
import re
import subprocess
import json
from datetime import datetime
from PyQt5.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QLineEdit, QPushButton, QMessageBox, QFileDialog,
                             QTableWidget, QTableWidgetItem, QHeaderView, QDialog,
                             QTabWidget, QTextEdit, QProgressBar, QMenu, QAction,
                             QSplitter, QFormLayout, QGroupBox, QScrollArea, QAbstractItemView,
                             QComboBox)
from PyQt5.QtCore import Qt, QSettings, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QFont, QIcon, QPalette, QColor, QCursor
from PyQt5 import QtCore
import openpyxl
from docxtpl import DocxTemplate

from widgets import ValidatedLineEdit, EditRecordDialog, RecordsTable
from update_manager import UpdateManager
from license_manager import LicenseManager


class DocumentWorker(QThread):
    """–ü–æ—Ç–æ–∫ –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤"""
    progress = pyqtSignal(int)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, save_root, fields, application_path):
        super().__init__()
        self.save_root = save_root
        self.fields = fields
        self.application_path = application_path

    def run(self):
        try:
            created_files = []

            # –°–æ–∑–¥–∞–µ–º –ø–∞–ø–∫—É –¥–ª—è –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤
            folder_name = f"{self.fields.get('n', '')} {self.fields.get('fn', '')} {self.fields.get('mn', '')}".strip()
            folder_path = os.path.join(self.save_root, folder_name)
            os.makedirs(folder_path, exist_ok=True)

            # –ü—É—Ç—å –∫ —à–∞–±–ª–æ–Ω–∞–º
            templates_dir = self.application_path

            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —à–∞–±–ª–æ–Ω—ã
            template_files = [f for f in os.listdir(templates_dir) if f.endswith('.docx')]
            if not template_files:
                self.error.emit("–®–∞–±–ª–æ–Ω—ã –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤ –Ω–µ –Ω–∞–π–¥–µ–Ω—ã")
                return

            total = len(template_files)
            for i, template_file in enumerate(template_files):
                self.progress.emit(int((i / total) * 100))

                template_path = os.path.join(templates_dir, template_file)

                # –ó–∞–≥—Ä—É–∂–∞–µ–º —à–∞–±–ª–æ–Ω
                doc = DocxTemplate(template_path)

                # –ü–æ–¥–≥–æ—Ç–∞–≤–ª–∏–≤–∞–µ–º –∫–æ–Ω—Ç–µ–∫—Å—Ç - –∏—Å–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –¥–ª—è –ø—Ä–∞–≤–∏–ª—å–Ω–æ–π –ø–æ–¥—Å—Ç–∞–Ω–æ–≤–∫–∏ –§–ò–û
                context = self.fields.copy()
                context['current_date'] = datetime.now().strftime('%d.%m.%Y')

                # –î–æ–±–∞–≤–ª—è–µ–º –ø–æ–ª—è –≤ –≤–µ—Ä—Ö–Ω–µ–º —Ä–µ–≥–∏—Å—Ç—Ä–µ
                context.update({
                    'n_c': context.get('n', '').upper(),
                    'fn_c': context.get('fn', '').upper(),
                    'mn_c': context.get('mn', '').upper()
                })

                # –†–µ–Ω–¥–µ—Ä–∏–º –¥–æ–∫—É–º–µ–Ω—Ç
                doc.render(context)

                # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç
                output_file = os.path.join(folder_path, template_file)
                doc.save(output_file)
                created_files.append(output_file)

            self.progress.emit(100)
            self.finished.emit(created_files)

        except Exception as e:
            self.error.emit(str(e))


class MainWindow(QMainWindow):
    """–ì–ª–∞–≤–Ω–æ–µ –æ–∫–Ω–æ –ø—Ä–∏–ª–æ–∂–µ–Ω–∏—è"""

    def __init__(self, settings, theme_manager):
        super().__init__()
        self.settings = settings
        self.theme_manager = theme_manager
        self.fields = {}
        self.records_data = []
        self.is_licensed = False

        self.init_ui()
        self.load_settings()

        # –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è –º–µ–Ω–µ–¥–∂–µ—Ä–æ–≤
        self.update_manager = UpdateManager()
        self.license_manager = LicenseManager(self.get_script_dir())

        # –ü—Ä–æ–≤–µ—Ä–∫–∞ –ª–∏—Ü–µ–Ω–∑–∏–∏
        self.check_license_on_startup()
        QTimer.singleShot(5000, self.check_for_updates_on_startup)

    def get_script_dir(self):
        """–ü–æ–ª—É—á–∏—Ç—å –¥–∏—Ä–µ–∫—Ç–æ—Ä–∏—é —Å–∫—Ä–∏–ø—Ç–∞"""
        if getattr(sys, 'frozen', False):
            return os.path.dirname(sys.executable)
        else:
            return os.path.dirname(os.path.abspath(__file__))

    def init_ui(self):
        """–ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å–∞"""
        self.setWindowTitle("–ü—Ä–æ–≥—Ä–∞–º–º–∞ –∑–∞–ø–æ–ª–Ω–µ–Ω–∏—è —Å–æ–≥–ª–∞—Å–∏–π –∏ –ª–∏—á–Ω—ã—Ö –∫–∞—Ä—Ç–æ—á–µ–∫")
        self.setGeometry(100, 100, 1200, 800)

        # –¶–µ–Ω—Ç—Ä–∞–ª—å–Ω—ã–π –≤–∏–¥–∂–µ—Ç
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # –û—Å–Ω–æ–≤–Ω–æ–π layout
        layout = QVBoxLayout(central_widget)

        # –°–æ–∑–¥–∞–µ–º —Ç–∞–±—ã
        self.tab_widget = QTabWidget()
        self.tab_widget.setFont(QFont("Segoe UI", 14))
        self.tab_widget.currentChanged.connect(self.on_tab_changed)
        layout.addWidget(self.tab_widget)

        # –í–∫–ª–∞–¥–∫–∞ –≤–≤–æ–¥–∞ –¥–∞–Ω–Ω—ã—Ö
        input_tab = QWidget()
        self.tab_widget.addTab(input_tab, "–í–≤–æ–¥ –¥–∞–Ω–Ω—ã—Ö")
        self.setup_input_tab(input_tab)

        # –í–∫–ª–∞–¥–∫–∞ –∑–∞–ø–∏—Å–µ–π
        records_tab = QWidget()
        self.tab_widget.addTab(records_tab, "–°–æ—Ö—Ä–∞–Ω–µ–Ω–Ω—ã–µ –∞–Ω–∫–µ—Ç—ã")
        self.setup_records_tab(records_tab)

        # –í–∫–ª–∞–¥–∫–∞ –Ω–∞—Å—Ç—Ä–æ–µ–∫
        settings_tab = QWidget()
        self.tab_widget.addTab(settings_tab, "–ù–∞—Å—Ç—Ä–æ–π–∫–∏")
        self.setup_settings_tab(settings_tab)

        # –°–æ–∑–¥–∞–µ–º –º–µ–Ω—é
        self.create_menu()

    def on_tab_changed(self, index):
        """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ —Å–º–µ–Ω—ã –≤–∫–ª–∞–¥–∫–∏"""
        try:
            if index == 1 and hasattr(self, 'records_table'):
                print("–ü–µ—Ä–µ–∫–ª—é—á–∏–ª–∏—Å—å –Ω–∞ –≤–∫–ª–∞–¥–∫—É —Å —Ç–∞–±–ª–∏—Ü–µ–π, –∑–∞–≥—Ä—É–∂–∞–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ...")
                QTimer.singleShot(50, self.records_table.load_state)
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–º–µ–Ω–µ –≤–∫–ª–∞–¥–∫–∏: {e}")

    def setup_input_tab(self, parent):
        """–ù–∞—Å—Ç—Ä–æ–π–∫–∞ –≤–∫–ª–∞–¥–∫–∏ –≤–≤–æ–¥–∞ –¥–∞–Ω–Ω—ã—Ö"""
        layout = QVBoxLayout(parent)
        layout.setSpacing(8)
        layout.setContentsMargins(8, 8, 8, 8)

        # –ü–æ–ª—è –≤–≤–æ–¥–∞
        form_widget = QWidget()
        form_layout = QFormLayout(form_widget)
        form_layout.setSpacing(6)
        form_layout.setContentsMargins(5, 5, 5, 5)

        for key, label in self.get_field_keys():
            if key == 'cs':
                field = ValidatedLineEdit('cyrillic_upper', 1)
            elif key in ['cn', 'ps', 'pn']:
                max_len = 6 if key in ['cn', 'pn'] else 4
                field = ValidatedLineEdit('digits', max_len)
            elif key == 'di':
                field = ValidatedLineEdit('date', 10)
            else:
                field = QLineEdit()

            field.setFont(QFont("Segoe UI", 14))
            label_widget = QLabel(label + ":")
            label_widget.setFont(QFont("Segoe UI", 14))
            form_layout.addRow(label_widget, field)
            self.fields[key] = field

        layout.addWidget(form_widget)

        # –ü–∞–ø–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è
        path_layout = QHBoxLayout()
        path_layout.setSpacing(8)

        path_label = QLabel("–ü–∞–ø–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è:")
        path_label.setFont(QFont("Segoe UI", 13))
        path_layout.addWidget(path_label)

        self.save_path_edit = QLineEdit()
        self.save_path_edit.setFont(QFont("Segoe UI", 13))
        self.save_path_edit.setText(self.get_default_save_folder())
        path_layout.addWidget(self.save_path_edit)

        browse_btn = QPushButton("–í—ã–±—Ä–∞—Ç—å...")
        browse_btn.setFont(QFont("Segoe UI", 12))
        browse_btn.clicked.connect(self.choose_folder)
        path_layout.addWidget(browse_btn)

        layout.addLayout(path_layout)

        # –ö–Ω–æ–ø–∫–∏ –¥–µ–π—Å—Ç–≤–∏–π
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)

        create_btn = QPushButton("–°–æ–∑–¥–∞—Ç—å –¥–æ–∫—É–º–µ–Ω—Ç—ã (–≤—Å–µ —à–∞–±–ª–æ–Ω—ã)")
        create_btn.clicked.connect(self.create_documents)
        create_btn.setFont(QFont("Segoe UI", 13))
        create_btn.setStyleSheet("QPushButton { padding: 12px; }")
        buttons_layout.addWidget(create_btn)

        save_btn = QPushButton("–°–æ—Ö—Ä–∞–Ω–∏—Ç—å –¥–∞–Ω–Ω—ã–µ")
        save_btn.clicked.connect(self.save_data)
        save_btn.setFont(QFont("Segoe UI", 13))
        save_btn.setStyleSheet("QPushButton { padding: 12px; }")
        buttons_layout.addWidget(save_btn)

        excel_btn = QPushButton("–û—Ç–∫—Ä—ã—Ç—å Excel")
        excel_btn.clicked.connect(self.open_excel)
        excel_btn.setFont(QFont("Segoe UI", 13))
        excel_btn.setStyleSheet("QPushButton { padding: 12px; }")
        buttons_layout.addWidget(excel_btn)

        layout.addLayout(buttons_layout)

        # –ü—Ä–æ–≥—Ä–µ—Å—Å-–±–∞—Ä
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

    def setup_records_tab(self, parent):
        """–ù–∞—Å—Ç—Ä–æ–π–∫–∞ –≤–∫–ª–∞–¥–∫–∏ –∑–∞–ø–∏—Å–µ–π"""
        layout = QVBoxLayout(parent)

        # –ö–Ω–æ–ø–∫–∏ —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è
        buttons_layout = QHBoxLayout()

        refresh_btn = QPushButton("–û–±–Ω–æ–≤–∏—Ç—å")
        refresh_btn.setFont(QFont("Segoe UI", 14))
        refresh_btn.clicked.connect(self.load_records)
        buttons_layout.addWidget(refresh_btn)

        load_btn = QPushButton("–ó–∞–≥—Ä—É–∑–∏—Ç—å –≤ —Ñ–æ—Ä–º—É")
        load_btn.setFont(QFont("Segoe UI", 14))
        load_btn.clicked.connect(self.load_selected_record)
        buttons_layout.addWidget(load_btn)

        edit_btn = QPushButton("–ò–∑–º–µ–Ω–∏—Ç—å")
        edit_btn.setFont(QFont("Segoe UI", 14))
        edit_btn.clicked.connect(self.edit_selected_record)
        buttons_layout.addWidget(edit_btn)

        delete_btn = QPushButton("–£–¥–∞–ª–∏—Ç—å")
        delete_btn.setFont(QFont("Segoe UI", 14))
        delete_btn.clicked.connect(self.delete_selected_record)
        buttons_layout.addWidget(delete_btn)

        layout.addLayout(buttons_layout)

        # –¢–∞–±–ª–∏—Ü–∞ –∑–∞–ø–∏—Å–µ–π
        self.records_table = RecordsTable(self.settings)
        self.records_table.setColumnCount(len(self.get_field_keys()) + 1)
        headers = [label for _, label in self.get_field_keys()] + ["RowNum"]
        self.records_table.setHorizontalHeaderLabels(headers)

        font = QFont("Segoe UI", 13)
        self.records_table.horizontalHeader().setFont(font)
        self.records_table.setFont(font)

        self.records_table.setColumnHidden(len(self.get_field_keys()), True)
        self.records_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.records_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.records_table.customContextMenuRequested.connect(self.show_records_context_menu)
        self.records_table.doubleClicked.connect(self.load_selected_record_double_click)

        layout.addWidget(self.records_table)
        self.load_records()

    def setup_settings_tab(self, parent):
        """–ù–∞—Å—Ç—Ä–æ–π–∫–∞ –≤–∫–ª–∞–¥–∫–∏ –Ω–∞—Å—Ç—Ä–æ–µ–∫ - —É–ø—Ä–æ—â–µ–Ω–Ω–∞—è –≤–µ—Ä—Å–∏—è"""
        layout = QVBoxLayout(parent)

        # –ì—Ä—É–ø–ø–∞ —Ç–µ–º
        theme_group = QGroupBox("–¢–µ–º–∞ –æ—Ñ–æ—Ä–º–ª–µ–Ω–∏—è")
        theme_group.setFont(QFont("Segoe UI", 12))
        theme_layout = QHBoxLayout(theme_group)

        self.light_theme_btn = QPushButton("–°–≤–µ—Ç–ª–∞—è")
        self.light_theme_btn.setFont(QFont("Segoe UI", 12))
        self.light_theme_btn.clicked.connect(lambda: self.change_theme('light'))
        theme_layout.addWidget(self.light_theme_btn)

        self.dark_theme_btn = QPushButton("–¢–µ–º–Ω–∞—è")
        self.dark_theme_btn.setFont(QFont("Segoe UI", 12))
        self.dark_theme_btn.clicked.connect(lambda: self.change_theme('dark'))
        theme_layout.addWidget(self.dark_theme_btn)

        layout.addWidget(theme_group)

        # –ì—Ä—É–ø–ø–∞ –ª–∏—Ü–µ–Ω–∑–∏–∏
        license_group = QGroupBox("–õ–∏—Ü–µ–Ω–∑–∏—è")
        license_group.setFont(QFont("Segoe UI", 13))
        license_layout = QVBoxLayout(license_group)

        # –ò–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è –æ –ª–∏—Ü–µ–Ω–∑–∏–∏
        license_info_layout = QHBoxLayout()
        license_type_label = QLabel("–¢–∏–ø –ª–∏—Ü–µ–Ω–∑–∏–∏:")
        license_type_label.setFont(QFont("Segoe UI", 12))
        license_info_layout.addWidget(license_type_label)
        self.license_type_label = QLabel("–ù–µ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–∞")
        self.license_type_label.setFont(QFont("Segoe UI", 12))
        license_info_layout.addWidget(self.license_type_label)
        license_info_layout.addStretch()

        license_days_label = QLabel("–û—Å—Ç–∞–ª–æ—Å—å –¥–Ω–µ–π:")
        license_days_label.setFont(QFont("Segoe UI", 12))
        license_info_layout.addWidget(license_days_label)
        self.license_days_label = QLabel("0")
        self.license_days_label.setFont(QFont("Segoe UI", 12))
        license_info_layout.addWidget(self.license_days_label)

        license_layout.addLayout(license_info_layout)

        # –ü–æ–ª–µ –¥–ª—è –≤–≤–æ–¥–∞ –∫–ª—é—á–∞
        key_layout = QHBoxLayout()
        key_label = QLabel("–õ–∏—Ü–µ–Ω–∑–∏–æ–Ω–Ω—ã–π –∫–ª—é—á:")
        key_label.setFont(QFont("Segoe UI", 13))
        key_layout.addWidget(key_label)

        self.license_edit = QLineEdit()
        self.license_edit.setFont(QFont("Segoe UI", 13))
        self.license_edit.setPlaceholderText("–í–≤–µ–¥–∏—Ç–µ –ª–∏—Ü–µ–Ω–∑–∏–æ–Ω–Ω—ã–π –∫–ª—é—á")
        key_layout.addWidget(self.license_edit)

        license_layout.addLayout(key_layout)

        # –ö–Ω–æ–ø–∫–∏ –ª–∏—Ü–µ–Ω–∑–∏–∏
        license_buttons_layout = QHBoxLayout()

        activate_btn = QPushButton("–ê–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å")
        activate_btn.setFont(QFont("Segoe UI", 13))
        activate_btn.clicked.connect(self.activate_license)
        license_buttons_layout.addWidget(activate_btn)

        license_layout.addLayout(license_buttons_layout)

        # –°—Ç–∞—Ç—É—Å –ª–∏—Ü–µ–Ω–∑–∏–∏
        self.license_status_label = QLabel("–°—Ç–∞—Ç—É—Å: –ù–µ –ø—Ä–æ–≤–µ—Ä–µ–Ω–æ")
        self.license_status_label.setFont(QFont("Segoe UI", 13))
        license_layout.addWidget(self.license_status_label)

        layout.addWidget(license_group)

        # –ì—Ä—É–ø–ø–∞ –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏
        info_group = QGroupBox("–û –ø—Ä–æ–≥—Ä–∞–º–º–µ")
        info_group.setFont(QFont("Segoe UI", 13))
        info_layout = QVBoxLayout(info_group)

        about_text = QTextEdit()
        about_text.setReadOnly(True)
        about_text.setFont(QFont("Segoe UI", 12))
        about_text.setHtml(f"""<pre style="font-family: 'Courier New', background: #f0f0f0; padding: 10px; border-radius: 5px;">
 üë®‚Äçüíª –†–ê–ó–†–ê–ë–û–¢–ß–ò–ö
 üìõ –°—Ç—Ä–æ—á–∫–æ–≤ –°–µ—Ä–≥–µ–π –ö–æ–Ω—Å—Ç–∞–Ω—Ç–∏–Ω–æ–≤–∏—á
 üìû 8(920)791-30-43
 üí¨ WhatsApp ‚Ä¢ Telegram
</pre>
        """)
        info_layout.addWidget(about_text)

        layout.addWidget(info_group)

        layout.addStretch()

        # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ª–∏—Ü–µ–Ω–∑–∏–∏ –ø–æ—Å–ª–µ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏–∏
        QTimer.singleShot(100, self.update_license_status)

    def create_menu(self):
        """–°–æ–∑–¥–∞–Ω–∏–µ –º–µ–Ω—é"""
        menubar = self.menuBar()
        menubar.setFont(QFont("Segoe UI", 12))

        # –ú–µ–Ω—é –§–∞–π–ª
        file_menu = menubar.addMenu('–§–∞–π–ª')

        save_action = QAction('–°–æ—Ö—Ä–∞–Ω–∏—Ç—å –¥–∞–Ω–Ω—ã–µ', self)
        save_action.setFont(QFont("Segoe UI", 12))
        save_action.triggered.connect(self.save_data)
        file_menu.addAction(save_action)

        create_action = QAction('–°–æ–∑–¥–∞—Ç—å –¥–æ–∫—É–º–µ–Ω—Ç—ã', self)
        create_action.setFont(QFont("Segoe UI", 14))
        create_action.triggered.connect(self.create_documents)
        file_menu.addAction(create_action)

        file_menu.addSeparator()

        exit_action = QAction('–í—ã—Ö–æ–¥', self)
        exit_action.setFont(QFont("Segoe UI", 12))
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # –ú–µ–Ω—é –í–∏–¥
        view_menu = menubar.addMenu('–í–∏–¥')

        light_theme_action = QAction('–°–≤–µ—Ç–ª–∞—è —Ç–µ–º–∞', self)
        light_theme_action.setFont(QFont("Segoe UI", 12))
        light_theme_action.triggered.connect(lambda: self.change_theme('light'))
        view_menu.addAction(light_theme_action)

        dark_theme_action = QAction('–¢–µ–º–Ω–∞—è —Ç–µ–º–∞', self)
        dark_theme_action.setFont(QFont("Segoe UI", 12))
        dark_theme_action.triggered.connect(lambda: self.change_theme('dark'))
        view_menu.addAction(dark_theme_action)

        # –ú–µ–Ω—é –°–µ—Ä–≤–∏—Å
        service_menu = menubar.addMenu('–°–µ—Ä–≤–∏—Å')

        update_action = QAction('–ü—Ä–æ–≤–µ—Ä–∏—Ç—å –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è', self)
        update_action.setFont(QFont("Segoe UI", 14))
        update_action.triggered.connect(self.check_for_updates)
        service_menu.addAction(update_action)

        service_menu.addSeparator()

        license_action = QAction('–ê–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å –ª–∏—Ü–µ–Ω–∑–∏—é', self)
        license_action.setFont(QFont("Segoe UI", 14))
        license_action.triggered.connect(self.show_license_dialog)
        service_menu.addAction(license_action)

        # –ú–µ–Ω—é –°–ø—Ä–∞–≤–∫–∞
        help_menu = menubar.addMenu('–°–ø—Ä–∞–≤–∫–∞')

        about_action = QAction('–û –ø—Ä–æ–≥—Ä–∞–º–º–µ', self)
        about_action.setFont(QFont("Segoe UI", 14))
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

    def get_field_keys(self):
        """–ü–æ–ª—É—á–∏—Ç—å –∫–ª—é—á–∏ –ø–æ–ª–µ–π"""
        return [
            ('n', '–§–∞–º–∏–ª–∏—è'),
            ('fn', '–ò–º—è'),
            ('mn', '–û—Ç—á–µ—Å—Ç–≤–æ'),
            ('reg', '–†–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏—è'),
            ('ps', '–°–µ—Ä–∏—è –ø–∞—Å–ø–æ—Ä—Ç–∞'),
            ('pn', '–ù–æ–º–µ—Ä –ø–∞—Å–ø–æ—Ä—Ç–∞'),
            ('pi', '–ü–∞—Å–ø–æ—Ä—Ç –≤—ã–¥–∞–Ω'),
            ('di', '–î–∞—Ç–∞ –≤—ã–¥–∞—á–∏'),
            ('cs', '–°–µ—Ä–∏—è –£–ß–û'),
            ('cn', '–ù–æ–º–µ—Ä –£–ß–û')
        ]

    def load_settings(self):
        """–ó–∞–≥—Ä—É–∑–∫–∞ –Ω–∞—Å—Ç—Ä–æ–µ–∫"""
        geometry = self.settings.get_window_geometry()
        if geometry:
            self.restoreGeometry(geometry)

        state = self.settings.get_window_state()
        if state:
            self.restoreState(state)

        last_path = self.settings.get_last_save_path()
        if last_path:
            self.save_path_edit.setText(last_path)

    def save_settings(self):
        """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–∞—Å—Ç—Ä–æ–µ–∫"""
        self.settings.set_window_geometry(self.saveGeometry())
        self.settings.set_window_state(self.saveState())
        self.settings.set_last_save_path(self.save_path_edit.text())

    def get_default_save_folder(self):
        """–ü–æ–ª—É—á–∏—Ç—å –ø–∞–ø–∫—É –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é"""
        return os.path.join(self.get_script_dir(), "–¥–æ–∫—É–º–µ–Ω—Ç—ã")

    def get_excel_file_path(self):
        """–ü–æ–ª—É—á–∏—Ç—å –ø—É—Ç—å –∫ Excel —Ñ–∞–π–ª—É"""
        return os.path.join(self.get_script_dir(), "–∞–Ω–∫–µ—Ç—ã_–¥–∞–Ω–Ω—ã–µ.xlsx")

    def ensure_excel_exists(self):
        """–°–æ–∑–¥–∞—Ç—å —Ñ–∞–π–ª Excel, –µ—Å–ª–∏ –æ–Ω –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç"""
        excel_path = self.get_excel_file_path()
        if not os.path.exists(excel_path):
            wb = openpyxl.Workbook()
            sheet = wb.active
            # –ó–∞–≥–æ–ª–æ–≤–∫–∏ —Å—Ç–æ–ª–±—Ü–æ–≤
            for col, (_, label) in enumerate(self.get_field_keys(), 1):
                sheet.cell(row=1, column=col, value=label)
            wb.save(excel_path)

    def get_field_values(self):
        """–ü–æ–ª—É—á–∏—Ç—å –∑–Ω–∞—á–µ–Ω–∏—è –≤—Å–µ—Ö –ø–æ–ª–µ–π"""
        values = {}
        for key, field in self.fields.items():
            values[key] = field.text().strip()
        return values

    def set_field_values(self, values):
        """–£—Å—Ç–∞–Ω–æ–≤–∏—Ç—å –∑–Ω–∞—á–µ–Ω–∏—è –ø–æ–ª–µ–π"""
        for key, value in values.items():
            if key in self.fields:
                self.fields[key].setText(value)

    def validate_fields(self, values, exclude_row=None):
        """–ü—Ä–æ–≤–µ—Ä–∫–∞ –ø–æ–ª–µ–π –Ω–∞ –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ—Å—Ç—å"""
        # –ü—Ä–æ–≤–µ—Ä–∫–∞ –æ–±—è–∑–∞—Ç–µ–ª—å–Ω—ã—Ö –ø–æ–ª–µ–π
        required_fields = ['n', 'fn', 'ps', 'pn', 'pi', 'di', 'cs', 'cn']
        for key in required_fields:
            if not values.get(key):
                field_name = dict(self.get_field_keys())[key]
                return False, f"–ü–æ–ª–µ '{field_name}' –æ–±—è–∑–∞—Ç–µ–ª—å–Ω–æ –¥–ª—è –∑–∞–ø–æ–ª–Ω–µ–Ω–∏—è."

        # –ü—Ä–æ–≤–µ—Ä–∫–∞ –ø–∞—Å–ø–æ—Ä—Ç–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö
        ps = values.get('ps', '')
        if not (ps.isdigit() and len(ps) == 4):
            return False, "–°–µ—Ä–∏—è –ø–∞—Å–ø–æ—Ä—Ç–∞ –¥–æ–ª–∂–Ω–∞ —Å–æ—Å—Ç–æ—è—Ç—å –∏–∑ 4 —Ü–∏—Ñ—Ä"

        pn = values.get('pn', '')
        if not (pn.isdigit() and len(pn) == 6):
            return False, "–ù–æ–º–µ—Ä –ø–∞—Å–ø–æ—Ä—Ç–∞ –¥–æ–ª–∂–µ–Ω —Å–æ—Å—Ç–æ—è—Ç—å –∏–∑ 6 —Ü–∏—Ñ—Ä"

        # –ü—Ä–æ–≤–µ—Ä–∫–∞ –£–ß–û
        cs = values.get('cs', '')
        if not re.match(r'^[–ê-–Ø–Å]$', cs):
            return False, "–°–µ—Ä–∏—è –£–ß–û –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –æ–¥–Ω–æ–π –∑–∞–≥–ª–∞–≤–Ω–æ–π —Ä—É—Å—Å–∫–æ–π –±—É–∫–≤–æ–π"

        cn = values.get('cn', '')
        if not (cn.isdigit() and len(cn) == 6):
            return False, "–ù–æ–º–µ—Ä –£–ß–û –¥–æ–ª–∂–µ–Ω —Å–æ—Å—Ç–æ—è—Ç—å –∏–∑ 6 —Ü–∏—Ñ—Ä"

        # –ü—Ä–æ–≤–µ—Ä–∫–∞ —É–Ω–∏–∫–∞–ª—å–Ω–æ—Å—Ç–∏ –£–ß–û
        if not self.is_cn_unique(cn, exclude_row):
            return False, "–ù–æ–º–µ—Ä –£–ß–û –¥–æ–ª–∂–µ–Ω –±—ã—Ç—å —É–Ω–∏–∫–∞–ª—å–Ω—ã–º"

        # –ü—Ä–æ–≤–µ—Ä–∫–∞ –¥–∞—Ç—ã
        di = values.get('di', '')
        if di:
            try:
                datetime.strptime(di, '%d.%m.%Y')
            except ValueError:
                return False, "–î–∞—Ç–∞ –≤—ã–¥–∞—á–∏ –ø–∞—Å–ø–æ—Ä—Ç–∞ –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –≤ —Ñ–æ—Ä–º–∞—Ç–µ –î–î.–ú–ú.–ì–ì–ì–ì"

        return True, ""

    def is_cn_unique(self, cn, exclude_row=None):
        """–ü—Ä–æ–≤–µ—Ä–∏—Ç—å —É–Ω–∏–∫–∞–ª—å–Ω–æ—Å—Ç—å –Ω–æ–º–µ—Ä–∞ –£–ß–û"""
        if not cn:
            return True

        excel_path = self.get_excel_file_path()
        if not os.path.exists(excel_path):
            return True

        try:
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active

            cn_col = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == "–ù–æ–º–µ—Ä –£–ß–û":
                    cn_col = col
                    break

            if cn_col is None:
                return True

            for row in range(2, sheet.max_row + 1):
                if exclude_row and row == exclude_row:
                    continue
                if str(sheet.cell(row=row, column=cn_col).value) == str(cn):
                    return False

            return True
        except Exception:
            return True

    def find_row_by_fullname(self, values):
        """–ù–∞–π—Ç–∏ —Å—Ç—Ä–æ–∫—É –ø–æ –§–ò–û"""
        excel_path = self.get_excel_file_path()
        if not os.path.exists(excel_path):
            return None

        try:
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active

            n_col, fn_col, mn_col = None, None, None
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header == "–§–∞–º–∏–ª–∏—è":
                    n_col = col
                elif header == "–ò–º—è":
                    fn_col = col
                elif header == "–û—Ç—á–µ—Å—Ç–≤–æ":
                    mn_col = col

            if n_col is None or fn_col is None:
                return None

            for row in range(2, sheet.max_row + 1):
                n_val = sheet.cell(row=row, column=n_col).value
                fn_val = sheet.cell(row=row, column=fn_col).value
                mn_val = sheet.cell(row=row, column=mn_col).value if mn_col else ""

                if (str(n_val) == str(values.get('n', '')) and
                        str(fn_val) == str(values.get('fn', '')) and
                        str(mn_val or '') == str(values.get('mn', ''))):
                    return row

            return None
        except Exception:
            return None

    def save_to_excel(self, values):
        """–°–æ—Ö—Ä–∞–Ω–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –≤ Excel"""
        try:
            excel_path = self.get_excel_file_path()
            self.ensure_excel_exists()

            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active

            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –ª–∏ —É–∂–µ –∑–∞–ø–∏—Å—å (–¥–ª—è –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è)
            existing_row = values.get('_row_number')

            if existing_row:
                # –û–±–Ω–æ–≤–ª—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â—É—é –∑–∞–ø–∏—Å—å
                for col, (key, _) in enumerate(self.get_field_keys(), 1):
                    sheet.cell(row=existing_row, column=col, value=values.get(key, ""))
                action = "–æ–±–Ω–æ–≤–ª–µ–Ω–∞"
            else:
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –ª–∏ –∑–∞–ø–∏—Å—å —Å —Ç–∞–∫–∏–º –§–ò–û
                existing_row_by_name = self.find_row_by_fullname(values)

                if existing_row_by_name:
                    # –û–±–Ω–æ–≤–ª—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â—É—é –∑–∞–ø–∏—Å—å
                    for col, (key, _) in enumerate(self.get_field_keys(), 1):
                        sheet.cell(row=existing_row_by_name, column=col, value=values.get(key, ""))
                    action = "–æ–±–Ω–æ–≤–ª–µ–Ω–∞"
                else:
                    # –î–æ–±–∞–≤–ª—è–µ–º –Ω–æ–≤—É—é –∑–∞–ø–∏—Å—å
                    row_num = sheet.max_row + 1
                    for col, (key, _) in enumerate(self.get_field_keys(), 1):
                        sheet.cell(row=row_num, column=col, value=values.get(key, ""))
                    action = "–¥–æ–±–∞–≤–ª–µ–Ω–∞"

            wb.save(excel_path)
            return True, f"–ê–Ω–∫–µ—Ç–∞ —É—Å–ø–µ—à–Ω–æ {action}."
        except Exception as e:
            return False, f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏: {str(e)}"

    def get_record_by_row_number(self, row_number):
        """–ù–∞–π—Ç–∏ –∑–∞–ø–∏—Å—å –ø–æ –Ω–æ–º–µ—Ä—É —Å—Ç—Ä–æ–∫–∏ –≤ Excel"""
        for record in self.records_data:
            if record.get('_row_number') == row_number:
                return record
        return None

    def get_selected_record_data(self):
        """–ü–æ–ª—É—á–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –≤—ã–±—Ä–∞–Ω–Ω–æ–π –∑–∞–ø–∏—Å–∏ —Å —É—á–µ—Ç–æ–º —Å–æ—Ä—Ç–∏—Ä–æ–≤–∫–∏"""
        try:
            selected_items = self.records_table.selectedItems()
            if not selected_items:
                return None

            # –ü–æ–ª—É—á–∞–µ–º –≤–∏–∑—É–∞–ª—å–Ω—É—é —Å—Ç—Ä–æ–∫—É
            visual_row = selected_items[0].row()

            # –ü–æ–ª—É—á–∞–µ–º –Ω–æ–º–µ—Ä —Å—Ç—Ä–æ–∫–∏ –≤ Excel –∏–∑ —Å–∫—Ä—ã—Ç–æ–π –∫–æ–ª–æ–Ω–∫–∏
            row_number_item = self.records_table.item(visual_row, len(self.get_field_keys()))
            if not row_number_item:
                return None

            row_number = int(row_number_item.text())
            return self.get_record_by_row_number(row_number)

        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –≤—ã–±—Ä–∞–Ω–Ω–æ–π –∑–∞–ø–∏—Å–∏: {e}")
            return None

    def load_records(self):
        """–ó–∞–≥—Ä—É–∑–∏—Ç—å –∑–∞–ø–∏—Å–∏ –≤ —Ç–∞–±–ª–∏—Ü—É"""
        try:
            excel_path = self.get_excel_file_path()
            if not os.path.exists(excel_path):
                self.ensure_excel_exists()
                return

            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active

            # –ü–æ–ª—É—á–∞–µ–º –¥–∞–Ω–Ω—ã–µ
            data = []
            for row in range(2, sheet.max_row + 1):
                record = {'_row_number': row}
                for col, (key, _) in enumerate(self.get_field_keys(), 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    record[key] = str(cell_value) if cell_value is not None else ""
                data.append(record)

            self.records_data = data

            # –í—Ä–µ–º–µ–Ω–Ω–æ –æ—Ç–∫–ª—é—á–∞–µ–º —Å–æ—Ä—Ç–∏—Ä–æ–≤–∫—É –¥–ª—è –∑–∞–ø–æ–ª–Ω–µ–Ω–∏—è
            self.records_table.setSortingEnabled(False)

            # –ó–∞–ø–æ–ª–Ω—è–µ–º —Ç–∞–±–ª–∏—Ü—É + –¥–æ–±–∞–≤–ª—è–µ–º —Å–∫—Ä—ã—Ç—É—é –∫–æ–ª–æ–Ω–∫—É —Å –Ω–æ–º–µ—Ä–æ–º —Å—Ç—Ä–æ–∫–∏
            self.records_table.setRowCount(len(data))
            for row_idx, record in enumerate(data):
                for col_idx, (key, _) in enumerate(self.get_field_keys()):
                    item = QTableWidgetItem(record.get(key, ""))
                    self.records_table.setItem(row_idx, col_idx, item)

                # –î–æ–±–∞–≤–ª—è–µ–º —Å–∫—Ä—ã—Ç—É—é –∫–æ–ª–æ–Ω–∫—É —Å –Ω–æ–º–µ—Ä–æ–º —Å—Ç—Ä–æ–∫–∏ –≤ Excel
                row_number_item = QTableWidgetItem(str(record['_row_number']))
                self.records_table.setItem(row_idx, len(self.get_field_keys()), row_number_item)

            # –í–∫–ª—é—á–∞–µ–º —Å–æ—Ä—Ç–∏—Ä–æ–≤–∫—É –æ–±—Ä–∞—Ç–Ω–æ
            self.records_table.setSortingEnabled(True)

            # –ü–æ—Å–ª–µ –∑–∞–≥—Ä—É–∑–∫–∏ –¥–∞–Ω–Ω—ã—Ö –∑–∞–≥—Ä—É–∂–∞–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ —Ç–∞–±–ª–∏—Ü—ã
            QTimer.singleShot(100, self.records_table.load_state)

        except Exception as e:
            QMessageBox.warning(self, "–û—à–∏–±–∫–∞", f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å –∑–∞–ø–∏—Å–∏: {str(e)}")

    def show_records_context_menu(self, position):
        """–ü–æ–∫–∞–∑–∞—Ç—å –∫–æ–Ω—Ç–µ–∫—Å—Ç–Ω–æ–µ –º–µ–Ω—é –¥–ª—è —Ç–∞–±–ª–∏—Ü—ã –∑–∞–ø–∏—Å–µ–π"""
        if not self.is_licensed:
            QMessageBox.warning(self, "–õ–∏—Ü–µ–Ω–∑–∏—è –Ω–µ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–∞",
                                "–î–ª—è —Ä–∞–±–æ—Ç—ã —Å –∑–∞–ø–∏—Å—è–º–∏ –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å –ª–∏—Ü–µ–Ω–∑–∏—é.")
            return

        try:
            menu = QMenu(self)

            load_action = menu.addAction("–ó–∞–≥—Ä—É–∑–∏—Ç—å –≤ —Ñ–æ—Ä–º—É")
            edit_action = menu.addAction("–ò–∑–º–µ–Ω–∏—Ç—å")
            delete_action = menu.addAction("–£–¥–∞–ª–∏—Ç—å")

            action = menu.exec_(self.records_table.viewport().mapToGlobal(position))

            if action == load_action:
                self.load_selected_record()
            elif action == edit_action:
                self.edit_selected_record()
            elif action == delete_action:
                self.delete_selected_record()
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –≤ –∫–æ–Ω—Ç–µ–∫—Å—Ç–Ω–æ–º –º–µ–Ω—é: {e}")

    def load_selected_record_double_click(self, index):
        """–ó–∞–≥—Ä—É–∑–∏—Ç—å –∑–∞–ø–∏—Å—å –ø–æ –¥–≤–æ–π–Ω–æ–º—É –∫–ª–∏–∫—É"""
        try:
            self.load_selected_record()
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –¥–≤–æ–π–Ω–æ–º –∫–ª–∏–∫–µ: {e}")

    def load_selected_record(self):
        """–ó–∞–≥—Ä—É–∑–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω—É—é –∑–∞–ø–∏—Å—å –≤ —Ñ–æ—Ä–º—É - —Å –ø—Ä–æ–≤–µ—Ä–∫–æ–π –ª–∏—Ü–µ–Ω–∑–∏–∏"""
        if not self.is_licensed:
            QMessageBox.warning(self, "–õ–∏—Ü–µ–Ω–∑–∏—è –Ω–µ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–∞",
                                "–î–ª—è –∑–∞–≥—Ä—É–∑–∫–∏ –∑–∞–ø–∏—Å–µ–π –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å –ª–∏—Ü–µ–Ω–∑–∏—é.")
            return

        try:
            record = self.get_selected_record_data()
            if not record:
                QMessageBox.warning(self, "–ù–µ –≤—ã–±—Ä–∞–Ω–æ", "–í—ã–±–µ—Ä–∏—Ç–µ –∑–∞–ø–∏—Å—å –¥–ª—è –∑–∞–≥—Ä—É–∑–∫–∏.")
                return

            self.set_field_values(record)
            QMessageBox.information(self, "–ì–æ—Ç–æ–≤–æ", "–î–∞–Ω–Ω—ã–µ –∑–∞–≥—Ä—É–∂–µ–Ω—ã –≤ —Ñ–æ—Ä–º—É.")
        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–≥—Ä—É–∑–∫–µ –∑–∞–ø–∏—Å–∏: {str(e)}")

    def edit_selected_record(self):
        """–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞—Ç—å –≤—ã–±—Ä–∞–Ω–Ω—É—é –∑–∞–ø–∏—Å—å - —Å –ø—Ä–æ–≤–µ—Ä–∫–æ–π –ª–∏—Ü–µ–Ω–∑–∏–∏"""
        if not self.is_licensed:
            QMessageBox.warning(self, "–õ–∏—Ü–µ–Ω–∑–∏—è –Ω–µ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–∞",
                                "–î–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –∑–∞–ø–∏—Å–µ–π –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å –ª–∏—Ü–µ–Ω–∑–∏—é.")
            return

        try:
            record = self.get_selected_record_data()
            if not record:
                QMessageBox.warning(self, "–ù–µ –≤—ã–±—Ä–∞–Ω–æ", "–í—ã–±–µ—Ä–∏—Ç–µ –∑–∞–ø–∏—Å—å –¥–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è.")
                return

            # –°–æ–∑–¥–∞–µ–º –∫–æ–ø–∏—é –¥–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è
            values = record.copy()

            # –°–æ–∑–¥–∞–µ–º –¥–∏–∞–ª–æ–≥
            dialog = EditRecordDialog(values, self)
            if dialog.exec_() == QDialog.Accepted:
                new_values = dialog.get_values()

                # –î–æ–±–∞–≤–ª—è–µ–º –Ω–æ–º–µ—Ä —Å—Ç—Ä–æ–∫–∏ –¥–ª—è –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è
                new_values['_row_number'] = record.get('_row_number')

                # –í–∞–ª–∏–¥–∞—Ü–∏—è
                valid, message = self.validate_fields(new_values, record.get('_row_number'))
                if not valid:
                    QMessageBox.warning(self, "–ù–µ–≤–µ—Ä–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ", message)
                    return

                # –°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ
                success, message = self.save_to_excel(new_values)
                if success:
                    QMessageBox.information(self, "–£—Å–ø–µ—Ö", message)
                    self.load_records()
                else:
                    QMessageBox.critical(self, "–û—à–∏–±–∫–∞", message)

        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–∏ –∑–∞–ø–∏—Å–∏: {str(e)}")

    def delete_selected_record(self):
        """–£–¥–∞–ª–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω—É—é –∑–∞–ø–∏—Å—å - —Å –ø—Ä–æ–≤–µ—Ä–∫–æ–π –ª–∏—Ü–µ–Ω–∑–∏–∏"""
        if not self.is_licensed:
            QMessageBox.warning(self, "–õ–∏—Ü–µ–Ω–∑–∏—è –Ω–µ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–∞",
                                "–î–ª—è —É–¥–∞–ª–µ–Ω–∏—è –∑–∞–ø–∏—Å–µ–π –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å –ª–∏—Ü–µ–Ω–∑–∏—é.")
            return

        try:
            record = self.get_selected_record_data()
            if not record:
                QMessageBox.warning(self, "–ù–µ –≤—ã–±—Ä–∞–Ω–æ", "–í—ã–±–µ—Ä–∏—Ç–µ –∑–∞–ø–∏—Å—å –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è.")
                return

            # –ü–æ–ª—É—á–∞–µ–º –§–ò–û –¥–ª—è –ø–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏—è
            fio = f"{record.get('n', '')} {record.get('fn', '')} {record.get('mn', '')}".strip()

            reply = QMessageBox.question(
                self,
                "–ü–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–µ —É–¥–∞–ª–µ–Ω–∏—è",
                f"–£–¥–∞–ª–∏—Ç—å –∑–∞–ø–∏—Å—å: {fio}?",
                QMessageBox.Yes | QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                try:
                    excel_path = self.get_excel_file_path()
                    wb = openpyxl.load_workbook(excel_path)
                    sheet = wb.active

                    # –£–¥–∞–ª—è–µ–º —Å—Ç—Ä–æ–∫—É (–∏—Å–ø–æ–ª—å–∑—É–µ–º —Å–æ—Ö—Ä–∞–Ω–µ–Ω–Ω—ã–π –Ω–æ–º–µ—Ä —Å—Ç—Ä–æ–∫–∏)
                    row_to_delete = record['_row_number']
                    sheet.delete_rows(row_to_delete)
                    wb.save(excel_path)

                    QMessageBox.information(self, "–£–¥–∞–ª–µ–Ω–æ", "–ó–∞–ø–∏—Å—å —É–¥–∞–ª–µ–Ω–∞.")
                    self.load_records()

                except Exception as e:
                    QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å –∑–∞–ø–∏—Å—å: {str(e)}")

        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —É–¥–∞–ª–µ–Ω–∏–∏ –∑–∞–ø–∏—Å–∏: {str(e)}")

    def choose_folder(self):
        """–í—ã–±–æ—Ä –ø–∞–ø–∫–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è"""
        try:
            path = QFileDialog.getExistingDirectory(
                self,
                "–í—ã–±–µ—Ä–∏—Ç–µ –ø–∞–ø–∫—É",
                self.save_path_edit.text() or self.get_default_save_folder()
            )
            if path:
                self.save_path_edit.setText(path)
                self.settings.set_last_save_path(path)
        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –≤—ã–±–æ—Ä–µ –ø–∞–ø–∫–∏: {str(e)}")

    def save_data(self):
        """–°–æ—Ö—Ä–∞–Ω–∏—Ç—å –¥–∞–Ω–Ω—ã–µ - —Å –ø—Ä–æ–≤–µ—Ä–∫–æ–π –ª–∏—Ü–µ–Ω–∑–∏–∏"""
        if not self.is_licensed:
            QMessageBox.warning(self, "–õ–∏—Ü–µ–Ω–∑–∏—è –Ω–µ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–∞",
                                "–î–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å –ª–∏—Ü–µ–Ω–∑–∏—é.")
            self.tab_widget.setCurrentIndex(2)
            return

        try:
            values = self.get_field_values()

            # –í–∞–ª–∏–¥–∞—Ü–∏—è
            existing_row = self.find_row_by_fullname(values)
            valid, message = self.validate_fields(values, existing_row)
            if not valid:
                QMessageBox.warning(self, "–ù–µ–≤–µ—Ä–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ", message)
                return

            # –°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ
            success, message = self.save_to_excel(values)
            if success:
                QMessageBox.information(self, "–£—Å–ø–µ—Ö", message)
                self.load_records()
            else:
                QMessageBox.critical(self, "–û—à–∏–±–∫–∞", message)
        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏ –¥–∞–Ω–Ω—ã—Ö: {str(e)}")

    def create_documents(self):
        """–°–æ–∑–¥–∞—Ç—å –¥–æ–∫—É–º–µ–Ω—Ç—ã - —Å –ø—Ä–æ–≤–µ—Ä–∫–æ–π –ª–∏—Ü–µ–Ω–∑–∏–∏"""
        if not self.is_licensed:
            QMessageBox.warning(self, "–õ–∏—Ü–µ–Ω–∑–∏—è –Ω–µ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–∞",
                                "–î–ª—è —Å–æ–∑–¥–∞–Ω–∏—è –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤ –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å –ª–∏—Ü–µ–Ω–∑–∏—é.")
            self.tab_widget.setCurrentIndex(2)
            return

        try:
            values = self.get_field_values()

            # –ü—Ä–æ–≤–µ—Ä–∫–∞ –æ–±—è–∑–∞—Ç–µ–ª—å–Ω—ã—Ö –ø–æ–ª–µ–π
            if not values.get('n') or not values.get('fn'):
                QMessageBox.warning(self, "–ü–æ–ª—è –Ω–µ –∑–∞–ø–æ–ª–Ω–µ–Ω—ã", "–§–∞–º–∏–ª–∏—è –∏ –∏–º—è –æ–±—è–∑–∞—Ç–µ–ª—å–Ω—ã.")
                return

            # –ü—Ä–æ–≤–µ—Ä–∫–∞ —Å—É—â–µ—Å—Ç–≤–æ–≤–∞–Ω–∏—è –∞–Ω–∫–µ—Ç—ã
            existing_row = self.find_row_by_fullname(values)
            if existing_row is None:
                reply = QMessageBox.question(
                    self,
                    "–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –∞–Ω–∫–µ—Ç—ã",
                    f"–ê–Ω–∫–µ—Ç–∞ –¥–ª—è {values.get('n')} {values.get('fn')} {values.get('mn', '')} –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ –≤ –±–∞–∑–µ.\n\n"
                    "–•–æ—Ç–∏—Ç–µ —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å –µ—ë –ø–µ—Ä–µ–¥ —Å–æ–∑–¥–∞–Ω–∏–µ–º –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply == QMessageBox.Yes:
                    success, message = self.save_to_excel(values)
                    if not success:
                        QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å –∞–Ω–∫–µ—Ç—É: {message}")
                        return

            # –ü—Ä–æ–≤–µ—Ä–∫–∞ –ø—É—Ç–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è
            save_root = self.save_path_edit.text() or self.get_default_save_folder()
            if not os.path.isdir(save_root):
                QMessageBox.critical(self, "–û—à–∏–±–∫–∞", "–ü—É—Ç—å —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–µ–Ω.")
                return

            # –ó–∞–ø—É—Å–∫ —Å–æ–∑–¥–∞–Ω–∏—è –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤ –≤ –æ—Ç–¥–µ–ª—å–Ω–æ–º –ø–æ—Ç–æ–∫–µ
            self.progress_bar.setVisible(True)
            self.worker = DocumentWorker(save_root, values, self.get_script_dir())
            self.worker.progress.connect(self.progress_bar.setValue)
            self.worker.finished.connect(self.on_documents_created)
            self.worker.error.connect(self.on_documents_error)
            self.worker.start()
        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤: {str(e)}")

    def on_documents_created(self, created_files):
        """–û–±—Ä–∞–±–æ—Ç–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∏—è —Å–æ–∑–¥–∞–Ω–∏—è –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤"""
        self.progress_bar.setVisible(False)

        if created_files:
            values = self.get_field_values()
            folder_name = f"{values.get('n', '')} {values.get('fn', '')} {values.get('mn', '')}".strip()

            QMessageBox.information(
                self,
                "–ì–æ—Ç–æ–≤–æ",
                f"–°–æ–∑–¥–∞–Ω–æ {len(created_files)} —Ñ–∞–π–ª–æ–≤:\n\n"
                f"–ü–∞–ø–∫–∞: {folder_name}\n"
                f"–ü—É—Ç—å: {os.path.join(self.save_path_edit.text(), folder_name)}"
            )
        else:
            QMessageBox.warning(
                self,
                "–í–Ω–∏–º–∞–Ω–∏–µ",
                "–î–æ–∫—É–º–µ–Ω—Ç—ã –Ω–µ —Å–æ–∑–¥–∞–Ω—ã. –ü—Ä–æ–≤–µ—Ä—å—Ç–µ –Ω–∞–ª–∏—á–∏–µ —à–∞–±–ª–æ–Ω–æ–≤ –∏ –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ—Å—Ç—å –¥–∞–Ω–Ω—ã—Ö."
            )

    def on_documents_error(self, error_message):
        """–û–±—Ä–∞–±–æ—Ç–∫–∞ –æ—à–∏–±–∫–∏ —Å–æ–∑–¥–∞–Ω–∏—è –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤"""
        self.progress_bar.setVisible(False)
        QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤: {error_message}")

    def open_excel(self):
        """–û—Ç–∫—Ä—ã—Ç—å —Ñ–∞–π–ª Excel - —Å –ø—Ä–æ–≤–µ—Ä–∫–æ–π –ª–∏—Ü–µ–Ω–∑–∏–∏"""
        if not self.is_licensed:
            QMessageBox.warning(self, "–õ–∏—Ü–µ–Ω–∑–∏—è –Ω–µ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–∞",
                                "–î–ª—è —Ä–∞–±–æ—Ç—ã —Å Excel –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å –ª–∏—Ü–µ–Ω–∑–∏—é.")
            self.tab_widget.setCurrentIndex(2)
            return

        try:
            excel_path = self.get_excel_file_path()
            if not os.path.exists(excel_path):
                self.ensure_excel_exists()

            if not os.path.exists(excel_path):
                QMessageBox.information(self, "–§–∞–π–ª –Ω–µ –Ω–∞–π–¥–µ–Ω", f"–§–∞–π–ª Excel –Ω–µ –Ω–∞–π–¥–µ–Ω –ø–æ –ø—É—Ç–∏: {excel_path}")
                return

            if sys.platform == "win32":
                os.startfile(excel_path)
            elif sys.platform == "darwin":
                subprocess.run(['open', excel_path])
            else:
                subprocess.run(['xdg-open', excel_path])
        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–ù–µ —É–¥–∞–ª–æ—Å—å –æ—Ç–∫—Ä—ã—Ç—å —Ñ–∞–π–ª: {e}")

    def change_theme(self, theme):
        """–ò–∑–º–µ–Ω–∏—Ç—å —Ç–µ–º—É –æ—Ñ–æ—Ä–º–ª–µ–Ω–∏—è"""
        try:
            self.theme_manager.apply_theme(theme)
            self.settings.set_theme(theme)
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–º–µ–Ω–µ —Ç–µ–º—ã: {e}")

    def check_for_updates(self):
        """–ü—Ä–æ–≤–µ—Ä–∏—Ç—å –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è - —É–ø—Ä–æ—â–µ–Ω–Ω–∞—è –≤–µ—Ä—Å–∏—è"""
        try:
            print("–ü—Ä–æ–≤–µ—Ä–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–π...")
            success, result = self.update_manager.check_for_updates()

            if success:
                if result == "up_to_date":
                    QMessageBox.information(self, "–û–±–Ω–æ–≤–ª–µ–Ω–∏—è",
                                            "‚úÖ –£—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∞ –ø–æ—Å–ª–µ–¥–Ω—è—è –≤–µ—Ä—Å–∏—è –ø—Ä–æ–≥—Ä–∞–º–º—ã.")
                else:
                    # –î–æ—Å—Ç—É–ø–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ
                    update_info = result
                    version = update_info.get('version', '–ù–æ–≤–∞—è –≤–µ—Ä—Å–∏—è')
                    download_url = update_info.get('download_url', '')

                    # –î–û–ü–û–õ–ù–ò–¢–ï–õ–¨–ù–ê–Ø –ü–†–û–í–ï–†–ö–ê URL
                    if not download_url.startswith('http'):
                        QMessageBox.warning(self, "–û—à–∏–±–∫–∞",
                                            f"–ù–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–π URL –¥–ª—è —Å–∫–∞—á–∏–≤–∞–Ω–∏—è: {download_url}")
                        return

                    reply = QMessageBox.question(
                        self,
                        "–î–æ—Å—Ç—É–ø–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ",
                        f"–î–æ—Å—Ç—É–ø–Ω–∞ –Ω–æ–≤–∞—è –≤–µ—Ä—Å–∏—è –ø—Ä–æ–≥—Ä–∞–º–º—ã: {version}\n\n"
                        f"URL: {download_url}\n\n"
                        "–£—Å—Ç–∞–Ω–æ–≤–∏—Ç—å –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ —Å–µ–π—á–∞—Å?",
                        QMessageBox.Yes | QMessageBox.No
                    )

                    if reply == QMessageBox.Yes:
                        self.install_update(update_info)
            else:
                QMessageBox.warning(self, "–û–±–Ω–æ–≤–ª–µ–Ω–∏—è",
                                    f"–ù–µ —É–¥–∞–ª–æ—Å—å –ø—Ä–æ–≤–µ—Ä–∏—Ç—å –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è:\n{result}")

        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞",
                                 f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–π:\n{str(e)}")

    def install_update(self, update_info):
        """–£—Å—Ç–∞–Ω–æ–≤–∏—Ç—å –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ"""
        try:
            reply = QMessageBox.question(
                self,
                "–ü–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–µ —É—Å—Ç–∞–Ω–æ–≤–∫–∏",
                "–ë—É–¥–µ—Ç —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ.\n\n"
                "–ü–µ—Ä–µ–¥ —É—Å—Ç–∞–Ω–æ–≤–∫–æ–π –±—É–¥–µ—Ç —Å–æ–∑–¥–∞–Ω–∞ —Ä–µ–∑–µ—Ä–≤–Ω–∞—è –∫–æ–ø–∏—è.\n"
                "–ü—Ä–æ–≥—Ä–∞–º–º–∞ –±—É–¥–µ—Ç –ø–µ—Ä–µ–∑–∞–ø—É—â–µ–Ω–∞ –ø–æ—Å–ª–µ —É—Å—Ç–∞–Ω–æ–≤–∫–∏.\n\n"
                "–ü—Ä–æ–¥–æ–ª–∂–∏—Ç—å?",
                QMessageBox.Yes | QMessageBox.No
            )

            if reply != QMessageBox.Yes:
                return

            # –°–æ–∑–¥–∞–µ–º –¥–∏–∞–ª–æ–≥ –ø—Ä–æ–≥—Ä–µ—Å—Å–∞
            progress_dialog = QMessageBox(self)
            progress_dialog.setWindowTitle("–£—Å—Ç–∞–Ω–æ–≤–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è")
            progress_dialog.setText("–í—ã–ø–æ–ª–Ω—è–µ—Ç—Å—è —É—Å—Ç–∞–Ω–æ–≤–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è...\n–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–¥–æ–∂–¥–∏—Ç–µ.")
            progress_dialog.setStandardButtons(QMessageBox.NoButton)
            progress_dialog.show()

            # –î–∞–µ–º –≤—Ä–µ–º—è –æ—Ç–æ–±—Ä–∞–∑–∏—Ç—å—Å—è –¥–∏–∞–ª–æ–≥—É
            QTimer.singleShot(100, lambda: self.perform_update_installation(update_info, progress_dialog))

        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —É—Å—Ç–∞–Ω–æ–≤–∫–µ –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è:\n{str(e)}")

    def perform_update_installation(self, update_info, progress_dialog):
        """–í—ã–ø–æ–ª–Ω–∏—Ç—å —É—Å—Ç–∞–Ω–æ–≤–∫—É –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è - –ò–°–ü–†–ê–í–õ–ï–ù–ù–´–ï –û–¢–°–¢–£–ü–´"""
        try:
            success, message = self.update_manager.download_and_install_update(update_info)
            progress_dialog.close()

            if success:
                QMessageBox.information(
                    self,
                    "–û–±–Ω–æ–≤–ª–µ–Ω–∏–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–æ",
                    "‚úÖ –û–±–Ω–æ–≤–ª–µ–Ω–∏–µ —É—Å–ø–µ—à–Ω–æ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–æ!\n\n"
                    "–ü—Ä–æ–≥—Ä–∞–º–º–∞ –±—É–¥–µ—Ç –ø–µ—Ä–µ–∑–∞–ø—É—â–µ–Ω–∞ –¥–ª—è –ø—Ä–∏–º–µ–Ω–µ–Ω–∏—è –∏–∑–º–µ–Ω–µ–Ω–∏–π."
                )
            else:
                QMessageBox.critical(
                    self,
                    "–û—à–∏–±–∫–∞ —É—Å—Ç–∞–Ω–æ–≤–∫–∏",
                    f"‚ùå –ù–µ —É–¥–∞–ª–æ—Å—å —É—Å—Ç–∞–Ω–æ–≤–∏—Ç—å –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ:\n{message}"
                )
        except Exception as e:
            progress_dialog.close()
            QMessageBox.critical(
                self,
                "–û—à–∏–±–∫–∞",
                f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ —É—Å—Ç–∞–Ω–æ–≤–∫–µ –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è:\n{str(e)}"
            )

    def activate_license(self):
        """–ê–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å –ª–∏—Ü–µ–Ω–∑–∏—é"""
        try:
            license_key = self.license_edit.text().strip()
            if not license_key:
                QMessageBox.warning(self, "–û—à–∏–±–∫–∞", "–í–≤–µ–¥–∏—Ç–µ –ª–∏—Ü–µ–Ω–∑–∏–æ–Ω–Ω—ã–π –∫–ª—é—á.")
                return

            success, message = self.license_manager.activate_license(license_key)
            if success:
                QMessageBox.information(self, "–£—Å–ø–µ—Ö", message)
                # –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ —Ä–∞–∑–±–ª–æ–∫–∏—Ä—É–µ–º –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å –ø–æ—Å–ª–µ —É—Å–ø–µ—à–Ω–æ–π –∞–∫—Ç–∏–≤–∞—Ü–∏–∏
                self.is_licensed = True
                self.unlock_interface()
                # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ª–∏—Ü–µ–Ω–∑–∏–∏
                self.update_license_status()
                # –û—á–∏—â–∞–µ–º –ø–æ–ª–µ –≤–≤–æ–¥–∞ –∫–ª—é—á–∞
                self.license_edit.clear()
            else:
                QMessageBox.critical(self, "–û—à–∏–±–∫–∞", message)
                # –û—Å—Ç–∞–≤–ª—è–µ–º –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω–Ω—ã–º
                self.is_licensed = False
                self.lock_interface()

            self.update_license_status()
        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∞–∫—Ç–∏–≤–∞—Ü–∏–∏ –ª–∏—Ü–µ–Ω–∑–∏–∏: {str(e)}")

    def show_license_dialog(self):
        """–ü–æ–∫–∞–∑–∞—Ç—å –¥–∏–∞–ª–æ–≥ –∞–∫—Ç–∏–≤–∞—Ü–∏–∏ –ª–∏—Ü–µ–Ω–∑–∏–∏"""
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("–ê–∫—Ç–∏–≤–∞—Ü–∏—è –ª–∏—Ü–µ–Ω–∑–∏–∏")
            dialog.setModal(True)
            dialog.resize(500, 250)

            layout = QVBoxLayout(dialog)

            info_label = QLabel(
                "–î–ª—è –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è –ø—Ä–æ–≥—Ä–∞–º–º—ã —Ç—Ä–µ–±—É–µ—Ç—Å—è –∞–∫—Ç–∏–≤–∞—Ü–∏—è –ª–∏—Ü–µ–Ω–∑–∏–∏.\n"
                "–ë–µ–∑ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–Ω–æ–π –ª–∏—Ü–µ–Ω–∑–∏–∏ –ø—Ä–æ–≥—Ä–∞–º–º–∞ –±—É–¥–µ—Ç –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω–∞."
            )
            info_label.setFont(QFont("Segoe UI", 12))
            info_label.setWordWrap(True)
            layout.addWidget(info_label)

            license_layout = QHBoxLayout()
            license_label = QLabel("–õ–∏—Ü–µ–Ω–∑–∏–æ–Ω–Ω—ã–π –∫–ª—é—á:")
            license_label.setFont(QFont("Segoe UI", 12))
            license_layout.addWidget(license_label)

            license_edit = QLineEdit()
            license_edit.setFont(QFont("Segoe UI", 12))
            license_edit.setPlaceholderText("–í–≤–µ–¥–∏—Ç–µ –ª–∏—Ü–µ–Ω–∑–∏–æ–Ω–Ω—ã–π –∫–ª—é—á...")
            license_layout.addWidget(license_edit)

            layout.addLayout(license_layout)

            buttons_layout = QHBoxLayout()
            activate_btn = QPushButton("–ê–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å")
            activate_btn.setFont(QFont("Segoe UI", 12))
            buttons_layout.addWidget(activate_btn)

            cancel_btn = QPushButton("–û—Ç–º–µ–Ω–∞")
            cancel_btn.setFont(QFont("Segoe UI", 12))
            buttons_layout.addWidget(cancel_btn)

            layout.addLayout(buttons_layout)

            status_label = QLabel("")
            status_label.setFont(QFont("Segoe UI", 11))
            status_label.setWordWrap(True)
            layout.addWidget(status_label)

            def activate():
                license_key = license_edit.text().strip()
                if not license_key:
                    status_label.setText("‚ùå –í–≤–µ–¥–∏—Ç–µ –ª–∏—Ü–µ–Ω–∑–∏–æ–Ω–Ω—ã–π –∫–ª—é—á.")
                    status_label.setStyleSheet("color: red;")
                    return

                success, message = self.license_manager.activate_license(license_key)
                if success:
                    status_label.setText("‚úÖ " + message)
                    status_label.setStyleSheet("color: green;")
                    # –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ —Ä–∞–∑–±–ª–æ–∫–∏—Ä—É–µ–º –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å
                    self.is_licensed = True
                    self.unlock_interface()
                    QTimer.singleShot(2000, dialog.accept)
                else:
                    status_label.setText("‚ùå " + message)
                    status_label.setStyleSheet("color: red;")

            activate_btn.clicked.connect(activate)
            cancel_btn.clicked.connect(dialog.reject)

            if dialog.exec_() == QDialog.Accepted:
                self.update_license_status()
        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–∫–∞–∑–µ –¥–∏–∞–ª–æ–≥–∞ –ª–∏—Ü–µ–Ω–∑–∏–∏: {str(e)}")

    def show_about(self):
        """–ü–æ–∫–∞–∑–∞—Ç—å –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –ø—Ä–æ–≥—Ä–∞–º–º–µ"""
        try:
            # –ò–º–ø–æ—Ä—Ç–∏—Ä—É–µ–º –≤–µ—Ä—Å–∏—é –∏–∑ version.py
            from version import __version__

            QMessageBox.about(
                self,
                "–û –ø—Ä–æ–≥—Ä–∞–º–º–µ",
                f"–ü—Ä–æ–≥—Ä–∞–º–º–∞ –∑–∞–ø–æ–ª–Ω–µ–Ω–∏—è —Å–æ–≥–ª–∞—Å–∏–π –∏ –ª–∏—á–Ω—ã—Ö –∫–∞—Ä—Ç–æ—á–µ–∫\n\n"
                f"–í–µ—Ä—Å–∏—è: {__version__}\n\n"
                "–†–∞–∑—Ä–∞–±–æ—Ç—á–∏–∫: –°—Ç—Ä–æ—á–∫–æ–≤ –°–µ—Ä–≥–µ–π –ö–æ–Ω—Å—Ç–∞–Ω—Ç–∏–Ω–æ–≤–∏—á\n"
                "–¢–µ–ª–µ—Ñ–æ–Ω: 8(920)791-30-43\n"
                "WhatsApp ‚Ä¢ Telegram"
            )
        except Exception as e:
            QMessageBox.critical(self, "–û—à–∏–±–∫–∞", f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–∫–∞–∑–µ –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏ –æ –ø—Ä–æ–≥—Ä–∞–º–º–µ: {str(e)}")

    def check_for_updates_on_startup(self):
        """–ü—Ä–æ–≤–µ—Ä–∏—Ç—å –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –ø—Ä–∏ –∑–∞–ø—É—Å–∫–µ - —Ç–∏—Ö–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞"""
        if hasattr(self, 'update_manager'):
            # –ó–∞–¥–µ—Ä–∂–∫–∞ —á—Ç–æ–±—ã –Ω–µ –º–µ—à–∞—Ç—å –∑–∞–ø—É—Å–∫—É
            QTimer.singleShot(5000, self.silent_update_check)

    def silent_update_check(self):
        """–¢–∏—Ö–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–π –±–µ–∑ –ø–æ–∫–∞–∑–∞ –¥–∏–∞–ª–æ–≥–æ–≤"""
        try:
            success, result = self.update_manager.check_for_updates()
            if success and result != "up_to_date":
                # –ü–æ–∫–∞–∑—ã–≤–∞–µ–º –Ω–µ–Ω–∞–≤—è–∑—á–∏–≤–æ–µ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ
                update_info = result
                version = update_info.get('version', '')
                if version.startswith('v'):
                    version = version[1:]

                # –°–æ–∑–¥–∞–µ–º –∫–∞—Å—Ç–æ–º–Ω–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ
                msg = QMessageBox(self)
                msg.setWindowTitle("–î–æ—Å—Ç—É–ø–Ω–æ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ")
                msg.setText(f"–î–æ—Å—Ç—É–ø–Ω–∞ –Ω–æ–≤–∞—è –≤–µ—Ä—Å–∏—è: {version}")
                msg.setInformativeText("–•–æ—Ç–∏—Ç–µ —É—Å—Ç–∞–Ω–æ–≤–∏—Ç—å –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ —Å–µ–π—á–∞—Å?")
                msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                msg.setDefaultButton(QMessageBox.No)

                reply = msg.exec_()
                if reply == QMessageBox.Yes:
                    self.install_update(update_info)
        except Exception as e:
            # –ò–≥–Ω–æ—Ä–∏—Ä—É–µ–º –æ—à–∏–±–∫–∏ –ø—Ä–∏ —Ç–∏—Ö–æ–π –ø—Ä–æ–≤–µ—Ä–∫–µ
            print(f"–¢–∏—Ö–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–π: {e}")

    def lock_interface(self):
        """–ó–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞—Ç—å –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å –ø—Ä–∏ –æ—Ç—Å—É—Ç—Å—Ç–≤–∏–∏ –ª–∏—Ü–µ–Ω–∑–∏–∏"""
        self.tab_widget.setTabEnabled(0, False)
        self.tab_widget.setTabEnabled(1, False)

        menubar = self.menuBar()
        for action in menubar.actions():
            if action.text() not in ['–°–µ—Ä–≤–∏—Å', '–°–ø—Ä–∞–≤–∫–∞']:
                action.setEnabled(False)

        self.light_theme_btn.setEnabled(False)
        self.dark_theme_btn.setEnabled(False)

        self.show_license_required_message()

    def unlock_interface(self):
        """–†–∞–∑–±–ª–æ–∫–∏—Ä–æ–≤–∞—Ç—å –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å –ø—Ä–∏ –Ω–∞–ª–∏—á–∏–∏ –ª–∏—Ü–µ–Ω–∑–∏–∏"""
        self.tab_widget.setTabEnabled(0, True)
        self.tab_widget.setTabEnabled(1, True)
        self.tab_widget.setTabEnabled(2, True)

        menubar = self.menuBar()
        for action in menubar.actions():
            action.setEnabled(True)

        self.light_theme_btn.setEnabled(True)
        self.dark_theme_btn.setEnabled(True)

        self.hide_license_required_message()

    def show_license_required_message(self):
        """–ü–æ–∫–∞–∑–∞—Ç—å —Å–æ–æ–±—â–µ–Ω–∏–µ –æ –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ—Å—Ç–∏ –∞–∫—Ç–∏–≤–∞—Ü–∏–∏"""
        if hasattr(self, 'license_message_label'):
            self.license_message_label.show()
            return

        self.license_message_label = QLabel(
            "‚ö†Ô∏è –¢–†–ï–ë–£–ï–¢–°–Ø –ê–ö–¢–ò–í–ê–¶–ò–Ø –õ–ò–¶–ï–ù–ó–ò–ò\n"
            "–î–ª—è –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è –ø—Ä–æ–≥—Ä–∞–º–º—ã –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å –ª–∏—Ü–µ–Ω–∑–∏—é –≤–æ –≤–∫–ª–∞–¥–∫–µ '–ù–∞—Å—Ç—Ä–æ–π–∫–∏'"
        )
        self.license_message_label.setAlignment(Qt.AlignCenter)
        self.license_message_label.setStyleSheet(
            "QLabel {"
            "background-color: #ffeb3b;"
            "color: #ff5722;"
            "font-weight: bold;"
            "font-size: 14px;"
            "padding: 10px;"
            "border: 2px solid #ff9800;"
            "border-radius: 5px;"
            "margin: 5px;"
            "}"
        )
        self.license_message_label.setFont(QFont("Segoe UI", 12, QFont.Bold))

        main_widget = self.centralWidget()
        main_layout = main_widget.layout()
        main_layout.insertWidget(0, self.license_message_label)

    def hide_license_required_message(self):
        """–°–∫—Ä—ã—Ç—å —Å–æ–æ–±—â–µ–Ω–∏–µ –æ –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ—Å—Ç–∏ –∞–∫—Ç–∏–≤–∞—Ü–∏–∏"""
        if hasattr(self, 'license_message_label'):
            self.license_message_label.hide()

    def check_license_on_startup(self):
        """–ü—Ä–æ–≤–µ—Ä–∏—Ç—å –ª–∏—Ü–µ–Ω–∑–∏—é –ø—Ä–∏ –∑–∞–ø—É—Å–∫–µ –ø—Ä–æ–≥—Ä–∞–º–º—ã"""
        print("–ü—Ä–æ–≤–µ—Ä–∫–∞ –ª–∏—Ü–µ–Ω–∑–∏–∏ –ø—Ä–∏ –∑–∞–ø—É—Å–∫–µ...")

        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ª–∏—Ü–µ–Ω–∑–∏—é
        license_check = self.license_manager.check_license()
        self.is_licensed = license_check[0]

        if not self.is_licensed:
            # –õ–∏—Ü–µ–Ω–∑–∏—è –Ω–µ –¥–µ–π—Å—Ç–≤–∏—Ç–µ–ª—å–Ω–∞ - –±–ª–æ–∫–∏—Ä—É–µ–º –ø—Ä–æ–≥—Ä–∞–º–º—É
            self.lock_interface()

            # –ü–æ–∫–∞–∑—ã–≤–∞–µ–º –∫—Ä–∏—Ç–∏—á–µ—Å–∫–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ
            QMessageBox.critical(
                self,
                "–õ–∏—Ü–µ–Ω–∑–∏—è –Ω–µ –¥–µ–π—Å—Ç–≤–∏—Ç–µ–ª—å–Ω–∞",
                f"–ü—Ä–æ–≥—Ä–∞–º–º–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –∑–∞–ø—É—â–µ–Ω–∞.\n\n–ü—Ä–∏—á–∏–Ω–∞: {license_check[2]}\n\n"
                "–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –∞–∫—Ç–∏–≤–∏—Ä—É–π—Ç–µ –ª–∏—Ü–µ–Ω–∑–∏—é –≤–æ –≤–∫–ª–∞–¥–∫–µ '–ù–∞—Å—Ç—Ä–æ–π–∫–∏'."
            )

            # –ü–µ—Ä–µ—Ö–æ–¥–∏–º –Ω–∞ –≤–∫–ª–∞–¥–∫—É –Ω–∞—Å—Ç—Ä–æ–µ–∫
            self.tab_widget.setCurrentIndex(2)
        else:
            # –õ–∏—Ü–µ–Ω–∑–∏—è –¥–µ–π—Å—Ç–≤–∏—Ç–µ–ª—å–Ω–∞ - —Ä–∞–∑–±–ª–æ–∫–∏—Ä—É–µ–º –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å
            self.unlock_interface()

        # –û–ë–ù–û–í–õ–Ø–ï–ú –°–¢–ê–¢–£–° –õ–ò–¶–ï–ù–ó–ò–ò –í –ò–ù–¢–ï–†–§–ï–ô–°–ï –ü–†–ò –ó–ê–ü–£–°–ö–ï
        self.update_license_status()

    def update_license_status(self):
        """–û–±–Ω–æ–≤–∏—Ç—å —Å—Ç–∞—Ç—É—Å –ª–∏—Ü–µ–Ω–∑–∏–∏ –≤ –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å–µ"""
        try:
            license_info = self.license_manager.get_license_info()

            is_valid = license_info['is_valid']
            days_left = license_info['days_left']
            message = license_info['message']
            license_type = license_info['type']
            is_trial = license_info.get('is_trial', False)

            if is_trial:
                self.license_type_label.setText("–ü—Ä–æ–±–Ω—ã–π –ø–µ—Ä–∏–æ–¥")
                status_text = f"–ü—Ä–æ–±–Ω—ã–π –ø–µ—Ä–∏–æ–¥ ({days_left} –¥–Ω–µ–π)"
            elif license_type == 'premium':
                self.license_type_label.setText("–ü—Ä–µ–º–∏—É–º")
                status_text = f"–ü—Ä–µ–º–∏—É–º ({days_left} –¥–Ω–µ–π)"
            else:
                self.license_type_label.setText("–ù–µ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–∞")
                status_text = "–ù–µ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–∞"

            self.license_days_label.setText(str(days_left))

            if is_valid:
                self.license_status_label.setText(f"–°—Ç–∞—Ç—É—Å: {status_text}")
            else:
                self.license_status_label.setText(f"–°—Ç–∞—Ç—É—Å: {message}")

            # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ –ª–∏—Ü–µ–Ω–∑–∏–∏ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö
            self.settings.settings.setValue("license/is_licensed", self.is_licensed)
            self.settings.settings.setValue("license/type", license_type)
            self.settings.settings.setValue("license/days_left", days_left)
            self.settings.settings.setValue("license/is_trial", is_trial)

        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–∏ —Å—Ç–∞—Ç—É—Å–∞ –ª–∏—Ü–µ–Ω–∑–∏–∏: {e}")

    def closeEvent(self, event):
        """–û–±—Ä–∞–±–æ—Ç–∫–∞ –∑–∞–∫—Ä—ã—Ç–∏—è –æ–∫–Ω–∞"""
        try:
            print("–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ —Å–æ—Å—Ç–æ—è–Ω–∏—è –ø—Ä–∏ –∑–∞–∫—Ä—ã—Ç–∏–∏ –ø—Ä–∏–ª–æ–∂–µ–Ω–∏—è...")

            # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ —Ç–∞–±–ª–∏—Ü—ã
            if hasattr(self, 'records_table'):
                print("–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ —Å–æ—Å—Ç–æ—è–Ω–∏—è —Ç–∞–±–ª–∏—Ü—ã...")
                self.records_table.save_state()

            # –°–æ—Ö—Ä–∞–Ω—è–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –ª–∏—Ü–µ–Ω–∑–∏–∏
            if hasattr(self, 'license_manager'):
                license_info = self.license_manager.get_license_info()
                self.settings.settings.setValue("license/is_licensed", license_info['is_valid'])
                self.settings.settings.setValue("license/type", license_info['type'])
                self.settings.settings.setValue("license/days_left", license_info['days_left'])
                self.settings.settings.setValue("license/is_trial", license_info.get('is_trial', False))

            self.save_settings()
            print("–ù–∞—Å—Ç—Ä–æ–π–∫–∏ —É—Å–ø–µ—à–Ω–æ —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã")
            event.accept()
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–∫—Ä—ã—Ç–∏–∏ –ø—Ä–∏–ª–æ–∂–µ–Ω–∏—è: {e}")
            event.accept()